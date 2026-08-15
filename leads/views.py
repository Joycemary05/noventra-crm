from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q, Count
from django.utils import timezone
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse
import openpyxl

from .models import Lead, FollowUp
from .forms import LeadForm


# =====================================
# LOGIN
# =====================================

def login_view(request):

    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user is not None:
            login(request, user)
            return redirect("dashboard")

        return render(
            request,
            "leads/login.html",
            {
                "error": "Invalid Username or Password"
            }
        )

    return render(request, "leads/login.html")


# =====================================
# LOGOUT
# =====================================

@login_required(login_url="login")
def logout_view(request):

    logout(request)

    return redirect("login")


# =====================================
# DASHBOARD
# =====================================

@login_required(login_url="login")
def dashboard(request):

    if request.user.is_superuser:
        leads = Lead.objects.all()
    else:
        leads = Lead.objects.filter(
            assigned_to=request.user
        )

    total_leads = leads.count()

    new_leads = leads.filter(status="new").count()
    contacted_leads = leads.filter(status="contacted").count()
    converted_leads = leads.filter(status="converted").count()

    recent_leads = leads.order_by("-created_at")[:5]

    source_stats = leads.values("source").annotate(
        total=Count("id")
    )

    if request.user.is_superuser:

        upcoming_followups = FollowUp.objects.filter(
            follow_up_date__gte=timezone.now().date()
        ).select_related("lead").order_by("follow_up_date")[:5]

    else:

        upcoming_followups = FollowUp.objects.filter(
            lead__assigned_to=request.user,
            follow_up_date__gte=timezone.now().date()
        ).select_related("lead").order_by("follow_up_date")[:5]

    status_labels = []
    status_data = []

    for item in leads.values("status").annotate(total=Count("id")):
        status_labels.append(item["status"].title())
        status_data.append(item["total"])

    source_labels = []
    source_data = []

    for item in source_stats:
        source_labels.append(item["source"].title())
        source_data.append(item["total"])

    return render(
        request,
        "leads/dashboard.html",
        {
            "total_leads": total_leads,
            "new_leads": new_leads,
            "contacted_leads": contacted_leads,
            "converted_leads": converted_leads,
            "recent_leads": recent_leads,
            "source_stats": source_stats,
            "upcoming_followups": upcoming_followups,
            "status_labels": status_labels,
            "status_data": status_data,
            "source_labels": source_labels,
            "source_data": source_data,
        }
    )


# =====================================
# ALL LEADS
# =====================================

@login_required(login_url="login")
def leads_list(request):

    search = request.GET.get("search", "")
    status = request.GET.get("status", "")

    if request.user.is_superuser:
        leads = Lead.objects.all().order_by("-created_at")
    else:
        leads = Lead.objects.filter(
            assigned_to=request.user
        ).order_by("-created_at")

    if search:
        leads = leads.filter(
            Q(name__icontains=search) |
            Q(email__icontains=search) |
            Q(company__icontains=search)
        )

    if status:
        leads = leads.filter(status=status)

    return render(
        request,
        "leads/leads_list.html",
        {
            "leads": leads,
            "search": search,
            "status": status,
        }
    )
# =====================================
# LEAD DETAILS
# =====================================

@login_required(login_url="login")
def lead_detail(request, lead_id):

    if request.user.is_superuser:
        lead = get_object_or_404(
            Lead,
            id=lead_id
        )
    else:
        lead = get_object_or_404(
            Lead,
            id=lead_id,
            assigned_to=request.user
        )

    followups = lead.followups.all().order_by("-follow_up_date")

    return render(
        request,
        "leads/lead_detail.html",
        {
            "lead": lead,
            "followups": followups,
        }
    )


# =====================================
# ADD LEAD
# =====================================

@login_required(login_url="login")
def add_lead(request):

    users = User.objects.filter(is_staff=True)

    if request.method == "POST":

        form = LeadForm(request.POST)

        if form.is_valid():

            lead = form.save(commit=False)

            if request.user.is_superuser:

                assigned_user = request.POST.get("assigned_to")

                if assigned_user:
                    lead.assigned_to = User.objects.get(id=assigned_user)

            else:

                lead.assigned_to = request.user

            lead.save()

            messages.success(
                request,
                "Lead added successfully!"
            )

            return redirect("leads_list")

    else:

        form = LeadForm()

    return render(
        request,
        "leads/add_lead.html",
        {
            "form": form,
            "users": users,
        }
    )


# =====================================
# EDIT LEAD
# =====================================

@login_required(login_url="login")
def edit_lead(request, lead_id):

    if request.user.is_superuser:

        lead = get_object_or_404(
            Lead,
            id=lead_id
        )

    else:

        lead = get_object_or_404(
            Lead,
            id=lead_id,
            assigned_to=request.user
        )

    if request.method == "POST":

        form = LeadForm(
            request.POST,
            instance=lead
        )

        if form.is_valid():

            lead = form.save(commit=False)

            if request.user.is_superuser:

                assigned_user = request.POST.get("assigned_to")

                if assigned_user:
                    lead.assigned_to = User.objects.get(id=assigned_user)
                else:
                    lead.assigned_to = None

            lead.save()

            messages.success(
                request,
                "Lead updated successfully!"
            )

            return redirect(
                "lead_detail",
                lead_id=lead.id
            )

    else:

        form = LeadForm(instance=lead)

    return render(
        request,
        "leads/edit_lead.html",
        {
            "form": form,
            "lead": lead,
        }
    )
# =====================================
# DELETE LEAD
# =====================================

@login_required(login_url="login")
def delete_lead(request, lead_id):

    if request.user.is_superuser:

        lead = get_object_or_404(
            Lead,
            id=lead_id
        )

    else:

        lead = get_object_or_404(
            Lead,
            id=lead_id,
            assigned_to=request.user
        )

    if request.method == "POST":

        lead.delete()

        messages.success(
            request,
            "Lead deleted successfully!"
        )

        return redirect("leads_list")

    return render(
        request,
        "leads/delete_lead.html",
        {
            "lead": lead
        }
    )


# =====================================
# ADD FOLLOW-UP
# =====================================

@login_required(login_url="login")
def add_followup(request, lead_id):

    if request.user.is_superuser:

        lead = get_object_or_404(
            Lead,
            id=lead_id
        )

    else:

        lead = get_object_or_404(
            Lead,
            id=lead_id,
            assigned_to=request.user
        )

    if request.method == "POST":

        FollowUp.objects.create(
            lead=lead,
            note=request.POST.get("note"),
            follow_up_date=request.POST.get("follow_up_date")
        )

        messages.success(
            request,
            "Follow-up added successfully!"
        )

        return redirect(
            "lead_detail",
            lead_id=lead.id
        )

    return render(
        request,
        "leads/add_followup.html",
        {
            "lead": lead
        }
    )


# =====================================
# IMPORT EXCEL
# =====================================

@login_required(login_url="login")
def import_excel(request):

    if not request.user.is_superuser:
        return redirect("dashboard")

    if request.method == "POST":

        excel_file = request.FILES.get("excel_file")

        if excel_file:

            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active

            for row in worksheet.iter_rows(min_row=2, values_only=True):

                Lead.objects.create(
                    name=row[0] or "",
                    email=row[1] or "",
                    phone=str(row[2] or ""),
                    company=row[3] or "",
                    source=(row[4] or "website").lower(),
                    status=(row[5] or "new").lower(),
                )

        messages.success(
            request,
            "Excel imported successfully!"
        )

        return redirect("leads_list")

    return render(
        request,
        "leads/import_excel.html"
    )


# =====================================
# EXPORT EXCEL
# =====================================

@login_required(login_url="login")
def export_excel(request):

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Leads"

    worksheet.append([
        "Name",
        "Email",
        "Phone",
        "Company",
        "Source",
        "Status",
        "Assigned To",
    ])

    if request.user.is_superuser:
        leads = Lead.objects.all()
    else:
        leads = Lead.objects.filter(
            assigned_to=request.user
        )

    for lead in leads:

        worksheet.append([
            lead.name,
            lead.email,
            lead.phone,
            lead.company,
            lead.get_source_display(),
            lead.get_status_display(),
            lead.assigned_to.username if lead.assigned_to else "Not Assigned",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="leads.xlsx"'
    )

    workbook.save(response)

    return response